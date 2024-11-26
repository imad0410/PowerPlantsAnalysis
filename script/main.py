import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, PatternFill
import datetime

# Constants
INPUT_FILE = '/mnt/data/conventional_power_plants_EU.csv'
OUTPUT_DIR = './output/'
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'PowerPlants_Report_{datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx')

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load and clean data
def load_and_clean_data(file_path):
    df = pd.read_csv(file_path)
    df.columns = df.columns.str.strip()  # Remove leading/trailing spaces in column names
    df = df.dropna(subset=['country', 'capacity', 'energy_source'])  # Drop rows missing critical information
    df['commissioned'] = pd.to_numeric(df['commissioned'], errors='coerce')
    return df

# Analyze data
def analyze_by_country(df):
    country_summary = df.groupby('country').agg(
        Total_Capacity=('capacity', 'sum'),
        Total_Power_Plants=('name', 'count')
    ).reset_index()
    return country_summary

def analyze_by_energy_source(df):
    energy_summary = df.groupby('energy_source').agg(
        Total_Capacity=('capacity', 'sum'),
        Total_Power_Plants=('name', 'count')
    ).reset_index()
    return energy_summary

def analyze_by_technology(df):
    technology_summary = df.groupby('technology').agg(
        Total_Capacity=('capacity', 'sum'),
        Total_Power_Plants=('name', 'count')
    ).reset_index()
    return technology_summary

# Generate Excel reports
def generate_excel_report(df, summaries):
    wb = Workbook()

    # Add main data sheet
    ws_data = wb.active
    ws_data.title = "All Power Plants"
    for row in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(row)

    # Style the sheet
    style_worksheet(ws_data, base_table_name="PowerPlantsTable")

    # Add summary sheets
    for sheet_name, summary_df in summaries.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in dataframe_to_rows(summary_df, index=False, header=True):
            ws.append(row)
        style_worksheet(ws, base_table_name=f"{sheet_name}Table")

    # Save the file
    wb.save(OUTPUT_FILE)
    print(f"Report generated: {OUTPUT_FILE}")

# Style Excel worksheet
def style_worksheet(ws, base_table_name):
    sanitized_table_name = base_table_name.replace(" ", "_")

    if ws.max_row > 1 and ws.max_column > 1:
        table_ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        table = Table(displayName=sanitized_table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Apply text alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

# Main script
if __name__ == "__main__":
    print("Loading and cleaning data...")
    data = load_and_clean_data(INPUT_FILE)

    print("Analyzing data...")
    country_summary = analyze_by_country(data)
    energy_summary = analyze_by_energy_source(data)
    technology_summary = analyze_by_technology(data)

    summaries = {
        "Summary by Country": country_summary,
        "Summary by Energy Source": energy_summary,
        "Summary by Technology": technology_summary
    }

    print("Generating Excel report...")
    generate_excel_report(data, summaries)

    print("Process complete.")
